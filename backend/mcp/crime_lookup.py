import json
from pathlib import Path

from openpyxl import load_workbook
from pyproj import Transformer
from shapely.geometry import Point, shape

DATA_DIR = Path(__file__).resolve().parent / "data"
CRIME_XLSX = DATA_DIR / "Fallzahlen&HZ 2016-2025.xlsx"
GEOJSON = DATA_DIR / "lor_bezirksregionen_2021.geojson"
YEAR = 2024
SKIP_CODES = {"999999"}  # citywide total row, not a real region

# Reverted: narrowing to violence-only columns (Raub, Straßenraub,
# Körperverletzungen, Freiheitsberaubung/Nötigung/Bedrohung) did NOT fix the
# low scores in high-foot-traffic tourist/transit hubs like Alexanderplatz —
# it made them slightly worse (39 -> 33). The real distortion is the HZ
# metric's denominator (cases per 100k *registered residents*, not visitors),
# which applies just as much to violent crime as to theft, since areas like
# Alexanderplatz have almost no actual residents. Switching categories can't
# fix a denominator problem, so back to the single "Straftaten insgesamt"
# column (index 2) — simplest, and no worse than the narrower attempt.
#
# FIX (this is the actual fix for the denominator problem above): the score
# is no longer normalized from HZ (cases per 100k residents). It's normalized
# from crimes-per-km2 (raw case count / Bezirksregion area), computed in
# attach_density() below. HZ is still loaded and kept on each entry for
# reference, but crime_safety_score comes from density. This stops
# near-zero-resident districts (Regierungsviertel, Alexanderplatz, Tiergarten)
# from reading as artificially dangerous just because almost nobody is
# officially registered as living there — while still flagging areas that
# are genuinely crime-dense per square km (e.g. Alexanderplatz remains
# elevated, correctly, since it really is a pickpocketing hotspot; but
# Tiergarten Süd — mostly parkland/embassies — jumps from 39 to 80).
CRIME_TOTAL_COL = 2  # "Straftaten -insgesamt-"

# lat/lng (EPSG:4326) -> the GeoJSON's native EPSG:25833 (UTM 33N, meters)
to_utm = Transformer.from_crs("EPSG:4326", "EPSG:25833", always_xy=True)

def load_crime_table():
    wb = load_workbook(CRIME_XLSX, read_only=True, data_only=True)
    fz_sheet = wb[f"Fallzahlen_{YEAR}"]
    hz_sheet = wb[f"HZ_{YEAR}"]

    def read_total(row):
        """Read the "Straftaten insgesamt" column for one row."""
        v = row[CRIME_TOTAL_COL]
        if v is None or v == "-":
            return None
        return float(v)

    fallzahl_by_code = {}
    for row in fz_sheet.iter_rows(min_row=6, values_only=True):
        code, name = row[0], row[1]
        if not code or code.endswith("0000") or code in SKIP_CODES:
            continue
        fallzahl_by_code[code] = {"name": name, "fallzahl": read_total(row)}

    result = {}
    for row in hz_sheet.iter_rows(min_row=6, values_only=True):
        code, name = row[0], row[1]
        if not code or code.endswith("0000") or code in SKIP_CODES:
            continue
        hz = read_total(row)
        if hz is None:
            continue
        entry = fallzahl_by_code.get(code, {"name": name, "fallzahl": None})
        entry["hz"] = hz
        result[code] = entry

    return result

def load_polygons():
    with open(GEOJSON, encoding="utf-8") as f:
        data = json.load(f)
    return [(feat["properties"]["bzr_id"], shape(feat["geometry"]))
            for feat in data["features"]]

def attach_density(crime_table, polygons):
    """Compute crimes-per-km2 for each region and store it as entry["density"].

    Polygons are already in EPSG:25833 (meters), so geom.area is m2 directly.
    Mutates and returns crime_table; regions with no fallzahl (raw count) or
    no matching polygon get density=None and are excluded from scoring range.
    """
    areas_km2 = {bzr_id: geom.area / 1_000_000 for bzr_id, geom in polygons}
    for code, entry in crime_table.items():
        fallzahl = entry.get("fallzahl")
        area_km2 = areas_km2.get(code)
        entry["density"] = (fallzahl / area_km2) if (fallzahl is not None and area_km2) else None
    return crime_table

def build_normalization(crime_table):
    densities = [v["density"] for v in crime_table.values() if v.get("density") is not None]
    return min(densities), max(densities)

def crime_score(lat, lng, crime_table, polygons, dens_min, dens_max):
    x, y = to_utm.transform(lng, lat)  # note: transform(x, y) = transform(lng, lat) with always_xy=True
    point = Point(x, y)
    for bzr_id, geom in polygons:
        if geom.contains(point):
            entry = crime_table.get(bzr_id)
            if not entry or entry.get("density") is None:
                return {"found": False, "region_code": bzr_id, "message": "matched region but no crime density for it"}
            pct = (entry["density"] - dens_min) / (dens_max - dens_min)
            pct = max(0.0, min(1.0, pct))
            score = round((1 - pct) * 100)
            return {
                "found": True,
                "region_code": bzr_id,
                "region_name": entry["name"],
                "crimes_per_km2": round(entry["density"], 1),
                "cases_per_100k": entry["hz"],  # kept for reference/debugging only; not used for scoring
                "crime_safety_score": score,
            }
    return {"found": False, "message": "point falls outside all Bezirksregionen"}

if __name__ == "__main__":
    crime_table = load_crime_table()
    polygons = load_polygons()
    attach_density(crime_table, polygons)
    dens_min, dens_max = build_normalization(crime_table)
    print(f"Loaded {len(crime_table)} regions. Density range: {dens_min:.1f} - {dens_max:.1f} crimes/km2")
    print(crime_score(52.5219, 13.4132, crime_table, polygons, dens_min, dens_max))